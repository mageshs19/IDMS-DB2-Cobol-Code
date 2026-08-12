import csv
import re
from io import BytesIO
from io import StringIO

from openpyxl import load_workbook

from idms_db2_phase2.domain.models import SheetMappingRow


class SheetMappingParser:
    COLUMNS = [
        "IDMS to DB2 Mapping",
        "Cobol Record IDMS",
        "Cobol Zone",
        "IDMS Key",
        "IDMS PIC Clause",
        "Length of Field Bytes",
        "Field end position",
        "DB2 Key",
        "New DB2 Record",
        "New DB2 Field name",
        "New DB2 Data Type",
        "Hopex Expression TypeRemark",
        "Remarks",
        "Relation",
        "Reference Field Name (CopyBook) ",
        "Reference Field PIC Clause",
        "Cross Application DB2 table",
        "Cross Application DB2 Field Name",
        "Cross Appln DB2 Data Type",
        "Basetype",
    ]

    FIELD_ALIASES = {
        "Cobol Record IDMS": [
            "Cobol Record IDMS",
            "COBOL RECORD IDMS",
            "Cobol Record",
            "COBOL Record",
            "IDMS Record",
            "Record IDMS",
            "Cobol Record Name",
            "IDMS to DB2 Mapping",

            # Production CSV variants seen in uploaded Sheet Mapping.
            "Cobol Recrd IDMS",
            "COBOL RECRD IDMS",
            "Cobol Rec IDMS",
            "COBOL REC IDMS",
            "Cobol Rcrd IDMS",
            "COBOL RCRD IDMS",
            "Cobol Recrd",
            "COBOL RECRD",
        ],
        "Cobol Zone": [
            "Cobol Zone",
            "COBOL Zone",
            "COBOL ZONE",
            "Cobol Field",
            "COBOL Field",
            "IDMS Field",
            "IDMS COBOL Zone",
            "IDMS COBOL Field",
            "Zone",
        ],
        "IDMS Key": [
            "IDMS Key",
            "IDMS KEY",
            "Key IDMS",
            "IDMS_Key",
        ],
        "IDMS PIC Clause": [
            "IDMS PIC Clause",
            "IDMS PIC",
            "PIC Clause",
            "Picture",
            "PIC",
        ],
        "Length of Field Bytes": [
            "Length of Field Bytes",
            "Length",
            "Field Length",
            "Length Bytes",
            "Length of Field",
        ],
        "Field end position": [
            "Field end position",
            "Field End Position",
            "End Position",
            "Field End",
        ],
        "DB2 Key": [
            "DB2 Key",
            "DB2 key",
            "DB2 KEY",
            "Key DB2",
            "DB2_key",
            "DB2_Key",
            "DB2KEY",
        ],
        "New DB2 Record": [
            "New DB2 Record",
            "New DB2 Record ",
            "DB2 Record",
            "DB2 Table",
            "New DB2 Table",
            "Table",
            "DB2_Table",
            "New DB2_Record",
        ],
        "New DB2 Field name": [
            "New DB2 Field name",
            "New DB2 Field Name",
            "New DB2_Field name",
            "New DB2_Field Name",
            "New DB2 Field",
            "New DB2_Field",
            "DB2 Field",
            "DB2 Column",
            "New DB2 Column",
            "Column",
            "DB2_Field",
            "DB2_Column",
        ],
        "New DB2 Data Type": [
            "New DB2 Data Type",
            "New DB2 DataType",
            "New DB2_Data Type",
            "New DB2_DataType",
            "DB2 Data Type",
            "DB2 DataType",
            "DB2 Type",
            "Data Type",
            "DataType",
        ],
        "Hopex Expression TypeRemark": [
            "Hopex Expression TypeRemark",
            "Hopex Expression Type Remark",
            "Hopex Expression Type",
            "Expression Type",
            "Expression Type Remark",
            "Hopex Type",
        ],
        "Remarks": [
            "Remarks",
            "Remark",
            "Comments",
            "Comment",
        ],
        "Relation": [
            "Relation",
            "Set",
            "Relationship",
            "Set Name",
            "IDMS Set",
            "IDMS Relation",
        ],
        "Reference Field Name (CopyBook) ": [
            "Reference Field Name (CopyBook) ",
            "Reference Field Name (CopyBook)",
            "Reference Field Name (Copybook)",
            "Reference Field Name CopyBook",
            "Reference Field Name Copybook",
            "Reference Field Name",
            "CopyBook Field",
            "Copybook Field",
            "Reference Field",
        ],
        "Reference Field PIC Clause": [
            "Reference Field PIC Clause",
            "Reference PIC",
            "Reference Field PIC",
            "Reference PIC Clause",
        ],
        "Cross Application DB2 table": [
            "Cross Application DB2 table",
            "Cross Application DB2 Table",
            "Cross App DB2 Table",
            "Cross Application Table",
            "Cross Application DB2 Record",
            "Cross Application DB2_Table",
        ],
        "Cross Application DB2 Field Name": [
            "Cross Application DB2 Field Name",
            "Cross App DB2 Field",
            "Cross Application Field Name",
            "Cross Application DB2 Column",
            "Cross Application DB2_Field Name",
            "Cross Application DB2_Field",
        ],
        "Cross Appln DB2 Data Type": [
            "Cross Appln DB2 Data Type",
            "Cross Appln DB2 DataType",
            "Cross Appln DB2_Data Type",
            "Cross Appln DB2_DataType",
            "Cross Application DB2 Data Type",
            "Cross Application DB2 DataType",
            "Cross App DB2 Type",
        ],
        "Basetype": [
            "Basetype",
            "Base Type",
            "BaseType",
        ],
    }

    def __init__(self) -> None:
        self.diagnostics: list[str] = []

    def parse_uploaded_file(
        self,
        uploaded_file,
    ) -> list[SheetMappingRow]:
        self.diagnostics = []

        if uploaded_file is None:
            self.diagnostics.append("No Sheet Mapping file supplied.")
            return []

        file_name = str(uploaded_file.name or "").lower()
        raw_bytes = uploaded_file.getvalue()

        self.diagnostics.append(f"Sheet Mapping file name: {file_name}")
        self.diagnostics.append(f"Sheet Mapping file size bytes: {len(raw_bytes)}")

        if file_name.endswith(".xlsx"):
            return self.parse_xlsx_bytes(raw_bytes)

        if file_name.endswith(".xls"):
            self.diagnostics.append(
                "Unsupported .xls file detected. Save the file as .xlsx or .csv."
            )
            return []

        text = raw_bytes.decode(
            "utf-8-sig",
            errors="ignore",
        )

        self.diagnostics.append(f"CSV/text decoded length: {len(text)}")

        if text:
            sample = text[:500].replace("\r", "\\r").replace("\n", "\\n")
            self.diagnostics.append(f"CSV/text sample: {sample}")

        return self.parse_csv_text(text)

    def parse_csv_text(
        self,
        text: str,
    ) -> list[SheetMappingRow]:
        if not str(text or "").strip():
            self.diagnostics.append("CSV/text Sheet Mapping is empty.")
            return []

        stream = StringIO(text)
        reader = csv.reader(stream)
        raw_rows = [tuple(row) for row in reader]

        if not raw_rows:
            self.diagnostics.append("CSV/text Sheet Mapping has no rows.")
            return []

        headers = [
            self._cell_to_string(value)
            for value in raw_rows[0]
        ]

        self.diagnostics.append(f"CSV detected headers: {headers}")

        output: list[SheetMappingRow] = []

        for row in raw_rows[1:]:
            raw_row: dict[str, str] = {}

            for index, header in enumerate(headers):
                if not header:
                    continue

                value = row[index] if index < len(row) else ""
                raw_row[header] = self._cell_to_string(value)

            normalized_raw_row = self._normalize_raw_dict_keys(raw_row)
            mapping_row = self._to_mapping_row(normalized_raw_row)

            if self._has_useful_content(mapping_row):
                output.append(mapping_row)

        self.diagnostics.append(f"CSV parsed useful rows: {len(output)}")
        self._add_population_diagnostics(output)

        return output

    def parse_xlsx_bytes(
        self,
        raw_bytes: bytes,
    ) -> list[SheetMappingRow]:
        if not raw_bytes:
            self.diagnostics.append("XLSX Sheet Mapping is empty.")
            return []

        workbook = load_workbook(
            BytesIO(raw_bytes),
            data_only=True,
            read_only=True,
        )

        output: list[SheetMappingRow] = []

        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                continue

            header_index = self._find_header_row(
                rows=rows,
                sheet_title=worksheet.title,
            )

            if header_index < 0:
                self.diagnostics.append(
                    f"Sheet {worksheet.title}: no Sheet Mapping header row detected."
                )
                continue

            parsed_rows = self._parse_xlsx_rows_from_header(
                raw_rows=rows,
                header_index=header_index,
            )

            self.diagnostics.append(
                f"Sheet {worksheet.title}: parsed useful rows: {len(parsed_rows)}"
            )

            output.extend(parsed_rows)

        self.diagnostics.append(f"XLSX parsed useful rows: {len(output)}")
        self._add_population_diagnostics(output)

        return output

    def _parse_xlsx_rows_from_header(
        self,
        raw_rows: list[tuple],
        header_index: int,
    ) -> list[SheetMappingRow]:
        headers = [
            self._cell_to_string(value)
            for value in raw_rows[header_index]
        ]

        output: list[SheetMappingRow] = []

        for row in raw_rows[header_index + 1:]:
            raw_row: dict[str, str] = {}

            for index, header in enumerate(headers):
                if not header:
                    continue

                value = row[index] if index < len(row) else ""
                raw_row[header] = self._cell_to_string(value)

            normalized_raw_row = self._normalize_raw_dict_keys(raw_row)
            mapping_row = self._to_mapping_row(normalized_raw_row)

            if self._has_useful_content(mapping_row):
                output.append(mapping_row)

        return output

    def _find_header_row(
        self,
        rows: list[tuple],
        sheet_title: str,
    ) -> int:
        canonical_groups = [
            ["Cobol Record IDMS", "Cobol Recrd IDMS", "IDMS Record"],
            ["New DB2 Record", "DB2 Table", "DB2 Record"],
            ["New DB2 Field name", "New DB2_Field name", "DB2 Column"],
        ]

        for index, row in enumerate(rows[:100]):
            normalized_cells = {
                self._normalize_header(self._cell_to_string(value))
                for value in row
                if value is not None
            }

            if index < 10:
                self.diagnostics.append(
                    f"Sheet {sheet_title}: row {index} normalized cells: "
                    f"{sorted(normalized_cells)}"
                )

            if self._normalize_header("IDMS to DB2 Mapping") in normalized_cells:
                return index

            if self._row_has_any_header(normalized_cells, canonical_groups[0]):
                return index

            if (
                self._row_has_any_header(normalized_cells, canonical_groups[1])
                and self._row_has_any_header(normalized_cells, canonical_groups[2])
            ):
                return index

        return -1

    def _row_has_any_header(
        self,
        normalized_cells: set[str],
        aliases: list[str],
    ) -> bool:
        for alias in aliases:
            if self._normalize_header(alias) in normalized_cells:
                return True

        return False

    def _to_mapping_row(
        self,
        raw_row: dict[str, str],
    ) -> SheetMappingRow:
        return SheetMappingRow(
            cobol_record_idms=self._get(
                raw_row,
                "Cobol Record IDMS",
            ),
            cobol_zone=self._get(
                raw_row,
                "Cobol Zone",
            ),
            idms_key=self._get(
                raw_row,
                "IDMS Key",
            ),
            idms_pic_clause=self._get(
                raw_row,
                "IDMS PIC Clause",
            ),
            length_of_field_bytes=self._get(
                raw_row,
                "Length of Field Bytes",
            ),
            field_end_position=self._get(
                raw_row,
                "Field end position",
            ),
            db2_key=self._get(
                raw_row,
                "DB2 Key",
            ),
            new_db2_record=self._get(
                raw_row,
                "New DB2 Record",
            ),
            new_db2_field_name=self._get(
                raw_row,
                "New DB2 Field name",
            ),
            new_db2_data_type=self._get(
                raw_row,
                "New DB2 Data Type",
            ),
            hopex_expression_type_remark=self._get(
                raw_row,
                "Hopex Expression TypeRemark",
            ),
            remarks=self._get(
                raw_row,
                "Remarks",
            ),
            relation=self._get(
                raw_row,
                "Relation",
            ),
            reference_field_name_copybook=self._get(
                raw_row,
                "Reference Field Name (CopyBook) ",
            ),
            reference_field_pic_clause=self._get(
                raw_row,
                "Reference Field PIC Clause",
            ),
            cross_application_db2_table=self._get(
                raw_row,
                "Cross Application DB2 table",
            ),
            cross_application_db2_field_name=self._get(
                raw_row,
                "Cross Application DB2 Field Name",
            ),
            cross_application_db2_data_type=self._get(
                raw_row,
                "Cross Appln DB2 Data Type",
            ),
            basetype=self._get(
                raw_row,
                "Basetype",
            ),
        )

    def _get(
        self,
        row: dict[str, str],
        canonical_name: str,
    ) -> str:
        aliases = self.FIELD_ALIASES.get(
            canonical_name,
            [canonical_name],
        )

        normalized_lookup = {
            self._normalize_header(key): value
            for key, value in row.items()
        }

        for alias in aliases:
            normalized_alias = self._normalize_header(alias)
            value = normalized_lookup.get(normalized_alias)

            if value is not None:
                return str(value).strip()

        return ""

    def _normalize_raw_dict_keys(
        self,
        row: dict,
    ) -> dict[str, str]:
        normalized: dict[str, str] = {}

        for key, value in row.items():
            clean_key = self._cell_to_string(key)
            clean_value = self._cell_to_string(value)

            if clean_key:
                normalized[clean_key] = clean_value

        return normalized

    def _cell_to_string(
        self,
        value,
    ) -> str:
        if value is None:
            return ""

        text = str(value)
        text = text.replace("\ufeff", "")
        text = text.replace("\xa0", " ")
        text = text.replace("\r\n", "\n")
        text = text.replace("\r", "\n")
        text = re.sub(r"[ \t]+", " ", text)

        return text.strip()

    def _normalize_header(
        self,
        value: str,
    ) -> str:
        text = self._cell_to_string(value)
        text = text.upper()
        text = text.replace("_", " ")
        text = re.sub(r"[^A-Z0-9]+", " ", text)
        text = re.sub(r"\s+", " ", text)

        return text.strip()

    def _has_useful_content(
        self,
        row: SheetMappingRow,
    ) -> bool:
        values = [
            row.cobol_record_idms,
            row.cobol_zone,
            row.idms_key,
            row.idms_pic_clause,
            row.db2_key,
            row.new_db2_record,
            row.new_db2_field_name,
            row.new_db2_data_type,
            row.relation,
            row.reference_field_name_copybook,
            row.cross_application_db2_table,
            row.cross_application_db2_field_name,
            row.basetype,
        ]

        return any(str(value or "").strip() for value in values)

    def _add_population_diagnostics(
        self,
        rows: list[SheetMappingRow],
    ) -> None:
        record_count = sum(
            1
            for row in rows
            if str(row.cobol_record_idms or "").strip()
        )

        source_field_count = sum(
            1
            for row in rows
            if str(row.cobol_zone or "").strip()
            or str(row.reference_field_name_copybook or "").strip()
        )

        db2_table_count = sum(
            1
            for row in rows
            if str(row.new_db2_record or "").strip()
            or str(row.cross_application_db2_table or "").strip()
        )

        db2_column_count = sum(
            1
            for row in rows
            if str(row.new_db2_field_name or "").strip()
            or str(row.cross_application_db2_field_name or "").strip()
        )

        useful_context_count = sum(
            1
            for row in rows
            if (
                str(row.cobol_record_idms or "").strip()
                and (
                    str(row.cobol_zone or "").strip()
                    or str(row.reference_field_name_copybook or "").strip()
                )
                and (
                    str(row.new_db2_record or "").strip()
                    or str(row.cross_application_db2_table or "").strip()
                )
                and (
                    str(row.new_db2_field_name or "").strip()
                    or str(row.cross_application_db2_field_name or "").strip()
                )
            )
        )

        self.diagnostics.append(
            f"Sheet Mapping populated Cobol Record IDMS rows: {record_count}"
        )
        self.diagnostics.append(
            f"Sheet Mapping populated source field rows: {source_field_count}"
        )
        self.diagnostics.append(
            f"Sheet Mapping populated DB2 table rows: {db2_table_count}"
        )
        self.diagnostics.append(
            f"Sheet Mapping populated DB2 column rows: {db2_column_count}"
        )
        self.diagnostics.append(
            f"Sheet Mapping useful source-to-target rows: {useful_context_count}"
        )